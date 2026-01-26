from django.test import TestCase, RequestFactory
from django.contrib.auth.models import User
from ..models import GameSession
from ..admin import export_to_xlsx, GameSessionAdmin
from django.contrib.admin.sites import AdminSite
import io
import openpyxl

class MockRequest:
    pass

class AdminExportTest(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.user = User.objects.create_user('user1', 'pass')
        self.session1 = GameSession.objects.create(user=self.user, score=100, level=1)
        self.session2 = GameSession.objects.create(user=self.user, score=200, level=2)
        self.site = AdminSite()
    #Тест функции выгрузки отчета в XLSX
    def test_export_to_xlsx(self):
        queryset = GameSession.objects.all()
        request = self.factory.get('/admin/')
        
        # Вызываем функцию напрямую
        response = export_to_xlsx(GameSessionAdmin(GameSession, self.site), request, queryset)
        
        # Проверяем заголовки ответа
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'], 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Читаем содержимое файла
        xlsx_content = io.BytesIO(response.content)
        workbook = openpyxl.load_workbook(xlsx_content)
        sheet = workbook.active
        
        # Проверяем заголовки таблицы (ID, User, Score, Level, Date)
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers, ['ID', 'User', 'Score', 'Level', 'Date'])
        
        # Проверяем данные первой строки (после заголовка)
        self.assertEqual(sheet.cell(row=2, column=2).value, 'user1') # User column
        self.assertEqual(sheet.cell(row=2, column=3).value, 100)     # Score column